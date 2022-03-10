import openpyxl


from qcloud_cos import CosConfig
from qcloud_cos import CosS3Client
from qcloud_cos import CosServiceError
from qcloud_cos import CosClientError

secret_id = 'xxxxxxxxxxxxx'     # 替换为用户的secret_id
secret_key = 'xxxxxxxxxxxx'     # 替换为用户的secret_key
region = 'ap-shanghai'    # 替换为用户的region
token = None               # 使用临时密钥需要传入Token，默认为空,可不填
config = CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key, Token=token)  # 获取配置对象
client = CosS3Client(config)
def up_tencent(localfilepath,key):
    client.upload_file(
        Bucket='huang-1258465420',
        LocalFilePath=localfilepath,#'D:\图片/v2-6cd79620efc0e75e5aed53df917addbd_r.jpg',
        Key=key,#'web_file/picture.jpg',
        PartSize=1,
        MAXThread=10,
        EnableMD5=False
)




import json

import random
import itertools
from flask import Flask,request,jsonify,render_template,redirect
#for i in list(itertools.permutations([1,2,3,4,5,6,7,8,9,10],10)):
    #result = random.sample(range(1,11),10)


app = Flask(__name__)


app.jinja_env.variable_start_string = '[['
app.jinja_env.variable_end_string = ']]'




tzidain =   {"bi" : 4,"bj" : 1,"bg" : 2,"bd" : 1,
    "ij" : 12,"ie" : 1,"ic" : 1,
    "hb" : 1,"hc" : 4,
    "gb" : 1,"gi" : 1,"gf" : 5,
    "fb" : 1,"fi" : 4,
    "ai" : 2,"ah" : 5,"ag" : 3,"ad" : 1,"ae" : 1,"ac" : 2,
    "db" : 1,"di" : 2,"dg" : 2,"dc" : 1,
    "eb" : 2,
    "cb" : 2,"ci" : 1,"cj" : 1,"cd" : 4,}

#-----次数结束-----






@app.route("/",methods = ['get','post'])
def q11():
    return render_template("money.html")




@app.route("/money1",methods = ['get','post'])
def qaz():
    try:
        txt1 = json.loads(request.form.get("data"))
        # print(txt1)
        txt = txt1["english"]

        canshu1 = txt

        canshu = canshu1.lower()
        qw = 1
        dict = {}
        for q in canshu:
            dict[q] = qw
            qw = qw + 1
        result = [dict['a'], dict['b'], dict['c'], dict['d'], dict['e'], dict['f'], dict['g'], dict['h'], dict['i'],
                  dict['j']]
        [a, b, c, d, e, f, g, h, i, j] = result

        def distance(q, w):

            return abs(dict[w] - dict[q])

        wbDataOnly2 = openpyxl.load_workbook(r'/data/www/excel/money.xlsx', data_only=False)

        #wbDataOnly2 = openpyxl.load_workbook("D:\桌面\money.xlsx", data_only=False)

        sheet1 = wbDataOnly2['Sheet1']

        # ------------------------------------------int---------

        # sheet1['A'].value = "次数表"

        hang_1 = 7

        lie_1 = 2

        zidian_hang_1 = {}

        zidian_lie_1 = {}

        for i in canshu1.upper():
            sheet1.cell(row=hang_1, column=1).value = i

            sheet1.cell(row=6, column=lie_1).value = i

            zidian_hang_1[i] = hang_1
            zidian_lie_1[i] = lie_1

            hang_1 = hang_1 + 1
            lie_1 = lie_1 + 1

        print(zidian_lie_1, zidian_hang_1)

        qqq_1 = 1
        for i in tzidain:
            hang_1 = zidian_hang_1[i[0].upper()]
            lie_1 = zidian_lie_1[i[1].upper()]
            sheet1.cell(row=hang_1, column=lie_1).value = tzidain[i]

            print("在第", qqq_1, "次", hang_1, lie_1, tzidain[i])
            # print(i)

            qqq_1 = qqq_1 + 1

        # --------------------------------

        sheet1 = wbDataOnly2['Sheet2']

        # sheet1['A'].value = "次数表"

        hang_2 = 7

        lie_2 = 2

        zidian_hang_2 = {}

        zidian_lie_2 = {}

        for i in canshu1.upper():
            sheet1.cell(row=hang_2, column=1).value = i

            sheet1.cell(row=6, column=lie_2).value = i

            zidian_hang_2[i] = hang_2
            zidian_lie_2[i] = lie_2

            hang_2 = hang_2 + 1
            lie_2 = lie_2 + 1

        qqq_2 = 1
        for i in tzidain:
            hang_2 = zidian_hang_2[i[0].upper()]
            lie_2 = zidian_lie_2[i[1].upper()]

            sheet1.cell(row=hang_2, column=lie_2).value = distance(i[0], i[1]) * tzidain[i]

            print("在第", qqq_2, "次", hang_2, lie_2, distance(i[0], i[1]) * tzidain[i])
            # print(i)

            qqq_2 = qqq_2 + 1

        # ------------------------------------------

        sheet1 = wbDataOnly2['Sheet3']

        # sheet1['A'].value = "次数表"

        hang = 7

        lie = 2

        zidian_hang = {}

        zidian_lie = {}
        #
        # for i in canshu1.upper():
        #     sheet1.cell(row=hang, column=0).value = i
        #
        #     sheet1.cell(row=6, column=lie).value = i
        #
        #     zidian_hang[i] = hang
        #     zidian_lie[i] = lie
        #
        #     hang = hang + 1
        #     lie = lie + 1

        qqq = 1

        for i in tzidain:

            lie = dict[i[1]]
            hang = dict[i[0]]

            if hang > lie:

                if sheet1.cell(row=lie, column=hang).value:

                    sheet1.cell(row=lie, column=hang).value = int(sheet1.cell(row=lie, column=hang).value) + distance(
                        i[0],
                        i[
                            1]) * \
                                                              tzidain[i]

                else:
                    sheet1.cell(row=lie, column=hang).value = distance(i[0], i[1]) * tzidain[i]

            else:

                if sheet1.cell(row=hang, column=lie).value:

                    sheet1.cell(row=hang, column=lie).value = int(sheet1.cell(row=hang, column=lie).value) + distance(
                        i[0],
                        i[
                            1]) * \
                                                              tzidain[i]

                else:
                    sheet1.cell(row=hang, column=lie).value = distance(i[0], i[1]) * tzidain[i]

            #  sheet1.cell(row=hang, column=lie).value = distance(i[0], i[1]) * tzidain[i]

            print("在第", qqq, "次", hang, lie, distance(i[0], i[1]) * tzidain[i])
            # print(i)

            qqq = qqq + 1

        wbDataOnly2.save(r'/data/www/excel/money/'+canshu+'.xlsx')

        excel_path = r'/data/www/excel/money/'+canshu+'.xlsx'
        #wbDataOnly2.save(r"D:/桌面/测试/" + canshu + '.xlsx')

        wbDataOnly2.close()

        up_tencent(excel_path, "web_file/picture/" + canshu+'.xlsx')

        return jsonify(canshu)

    except:
        return jsonify("格式错误")



if __name__ == '__main__':

    app.run(host='127.0.0.1', port=5012, debug='True')


